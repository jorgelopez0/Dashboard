import os
import shutil
from pathlib import Path

import openpyxl
from datetime import datetime
from collections import defaultdict

EXCEL_PATH = os.getenv('EXCEL_PATH', 'cash.xlsx')
BACKUP_DIR = Path(EXCEL_PATH).parent / 'excel_backups'
MAX_BACKUPS = 50

# En Render (free tier, sin disco persistente) EXCEL_PATH vive en el filesystem
# efímero del contenedor. EXCEL_SEED_PATH puede apuntar a un Secret File de solo
# lectura con una copia real de cash.xlsx; al arrancar se clona a EXCEL_PATH para
# que la app tenga datos reales y pueda escribir hasta el próximo reinicio/deploy.
_seed_path = os.getenv('EXCEL_SEED_PATH')
if _seed_path and os.path.exists(_seed_path) and not os.path.exists(EXCEL_PATH):
    shutil.copy2(_seed_path, EXCEL_PATH)

BOOKMAKERS = ['WINAMAX', 'CODERE', 'BET365', 'GCM', 'BETFAIR', 'BWIN',
              'VERSUS', 'BETWAY', 'YOSPORTS', 'OLYBET', 'WILLIAMHILL', 'CASINOMAR BELLA']
SHARED_BOOKMAKERS = {'WINAMAX', 'CODERE', 'BET365', 'GCM', 'BETWAY', 'VERSUS', 'YOSPORTS', 'BWIN'}
SALDOS_COLS = ['EFECTIVO', 'REVOLUT'] + BOOKMAKERS + ['CLASES', 'PAPIS', 'TRIBBU', 'ENTRENADOR']
TIPOS_MOV = ['Retirada', 'Ingreso', 'Gasto', 'Papis', 'Clases', 'Tribbu', 'Entrenador']
TIPOS_MOV_SIMPLES = {'Gasto', 'Papis', 'Clases', 'Tribbu', 'Entrenador'}  # sin cuenta/casa/estado: solo suman o restan de Revolut
CUENTAS = ['Yo', 'Asenjo']
CUENTA_LABELS = {'Yo': 'Cuenta 1', 'Asenjo': 'Cuenta 2'}
CUENTAS_APUESTA = ['Ambas', 'Yo', 'Asenjo']
CUENTAS_APUESTA_LABELS = {'Ambas': '1 y 2', 'Yo': '1', 'Asenjo': '2'}
CUENTAS_APUESTA_OPTION_LABELS = {'Ambas': 'Ambas', 'Yo': '1', 'Asenjo': '2'}
RESULTADOS = ['Pendiente', 'W', 'L', 'V']
MONTH_NAMES = ['ene', 'feb', 'mar', 'abr', 'may', 'jun',
               'jul', 'ago', 'sep', 'oct', 'nov', 'dic']


def _num(val):
    return float(val) if isinstance(val, (int, float)) else 0.0


def _fmt_mes(dt):
    if not dt:
        return ''
    if isinstance(dt, datetime):
        return f"{MONTH_NAMES[dt.month - 1]}-{str(dt.year)[2:]}"
    return str(dt)


def _backup_before_write():
    """Copia cash.xlsx a excel_backups/ antes de sobrescribirlo, por si un guardado
    lo corrompe o borra algo sin querer. Guarda como mucho MAX_BACKUPS copias."""
    try:
        if not os.path.exists(EXCEL_PATH):
            return
        BACKUP_DIR.mkdir(exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        shutil.copy2(EXCEL_PATH, BACKUP_DIR / f'cash_{ts}.xlsx')
        backups = sorted(BACKUP_DIR.glob('cash_*.xlsx'))
        for old in backups[:-MAX_BACKUPS]:
            old.unlink(missing_ok=True)
    except OSError:
        pass  # el backup nunca debe bloquear el guardado real


def _first_of_month(dt):
    if isinstance(dt, datetime):
        return datetime(dt.year, dt.month, 1)
    return dt


def _calc_ganancia(resultado, stake, cuota):
    if resultado == 'W':
        return round(stake * (cuota - 1), 2)
    elif resultado == 'L':
        return round(-stake, 2)
    elif resultado == 'V':
        return 0.0
    return None


def _normalize_cuenta(val):
    """Corrige diferencias de mayúsculas/minúsculas en el campo 'cuenta' de Movimientos
    (p.ej. 'yo' en vez de 'Yo') para que no se pierdan silenciosamente del cálculo de saldos."""
    if not val:
        return val
    for c in CUENTAS:
        if str(val).strip().lower() == c.lower():
            return c
    return val


def _normalize_cuentas(casa, cuentas):
    """Determina el reparto real de cuentas para una apuesta: solo las casas con
    cuenta compartida (SHARED_BOOKMAKERS) admiten elegir; el resto va siempre a Cuenta 1 (Yo)."""
    is_shared = casa in SHARED_BOOKMAKERS if casa else False
    if not is_shared:
        return 'Yo'
    return cuentas if cuentas in CUENTAS_APUESTA else 'Ambas'


def _split_ganancia(gan, cuentas_v):
    if gan is None:
        return None, None
    if cuentas_v == 'Ambas':
        return gan, gan
    elif cuentas_v == 'Yo':
        return gan, 0.0
    else:
        return 0.0, gan


def load_initial_balances():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Saldos']
    yo = {}
    asenjo = {}
    for i, col_name in enumerate(SALDOS_COLS):
        col_num = i + 2  # B=2
        yo[col_name] = _num(ws.cell(row=4, column=col_num).value)
        asenjo[col_name] = _num(ws.cell(row=8, column=col_num).value)
    return yo, asenjo


def load_movimientos():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Movimientos']
    result = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        fecha, mes, tipo, cuenta, casa, cantidad, estado = (list(row) + [None] * 7)[:7]
        if mes is None and fecha is None:
            continue
        result.append({
            'row': row_num,
            'fecha': fecha,
            'mes': mes,
            'mes_str': _fmt_mes(mes),
            'tipo': tipo,
            'cuenta': _normalize_cuenta(cuenta),
            'casa': casa or '',
            'cantidad': _num(cantidad),
            'estado': estado or 'Pendiente',
        })
    return result


def load_apuestas():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Registro_Apuestas']
    result = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row) + [None] * 12
        mes, fecha, partido, apuesta_desc, casa, stake, cuota, resultado = vals[:8]
        cuentas_raw = vals[11]
        if mes is None and casa is None:
            continue
        stake_v = _num(stake)
        cuota_v = _num(cuota)
        resultado_v = resultado if resultado else 'Pendiente'
        cuentas_v = _normalize_cuentas(casa, cuentas_raw)
        gan = _calc_ganancia(resultado_v, stake_v, cuota_v)
        gan_yo, gan_asenjo = _split_ganancia(gan, cuentas_v)
        pendiente = resultado_v == 'Pendiente'
        result.append({
            'row': row_num,
            'mes': mes,
            'mes_str': _fmt_mes(mes),
            'fecha': fecha,
            'partido': partido or '',
            'apuesta': apuesta_desc or '',
            'casa': casa or '',
            'stake': stake_v,
            'cuota': cuota_v,
            'resultado': resultado_v,
            'cuentas': cuentas_v,
            'is_shared': casa in SHARED_BOOKMAKERS if casa else False,
            'ganancia_yo': gan_yo if gan_yo is not None else 0.0,
            'ganancia_asenjo': gan_asenjo if gan_asenjo is not None else 0.0,
            'pendiente': pendiente,
        })
    return result


def compute_saldos(movimientos, apuestas, yo_init, asenjo_init):
    # REVOLUT
    revolut = yo_init['REVOLUT']
    for m in movimientos:
        if m['estado'] == 'Recibido':
            t = m['tipo']
            if t == 'Ingreso':
                revolut -= m['cantidad']
            elif t == 'Retirada':
                revolut += m['cantidad']
            elif t == 'Gasto':
                revolut -= m['cantidad']
            elif t in ('Papis', 'Clases', 'Tribbu', 'Entrenador'):
                revolut += m['cantidad']

    yo = {'EFECTIVO': yo_init['EFECTIVO'], 'REVOLUT': round(revolut, 2)}
    asenjo = {}

    for bm in BOOKMAKERS:
        sy = yo_init.get(bm, 0.0)
        sa = asenjo_init.get(bm, 0.0)
        for ap in apuestas:
            if ap['casa'] == bm and not ap['pendiente']:
                sy += ap['ganancia_yo']
                sa += ap['ganancia_asenjo']
        for m in movimientos:
            if m['casa'] == bm and m['tipo'] == 'Ingreso' and m['estado'] == 'Recibido':
                if m['cuenta'] == 'Yo':
                    sy += m['cantidad']
                elif m['cuenta'] == 'Asenjo':
                    sa += m['cantidad']
            if m['casa'] == bm and m['tipo'] == 'Retirada':
                if m['cuenta'] == 'Yo':
                    sy -= m['cantidad']
                elif m['cuenta'] == 'Asenjo':
                    sa -= m['cantidad']
        yo[bm] = round(sy, 2)
        asenjo[bm] = round(sa, 2)

    for cat, tipo_val in [('CLASES', 'Clases'), ('PAPIS', 'Papis'), ('TRIBBU', 'Tribbu'), ('ENTRENADOR', 'Entrenador')]:
        saldo = yo_init.get(cat, 0.0)
        for m in movimientos:
            if m['tipo'] == tipo_val and m['estado'] == 'Pendiente':
                saldo += m['cantidad']
        yo[cat] = round(saldo, 2)

    yo_apostado = defaultdict(float)
    asenjo_apostado = defaultdict(float)
    for ap in apuestas:
        if ap['pendiente'] and ap['casa']:
            if ap['cuentas'] in ('Ambas', 'Yo'):
                yo_apostado[ap['casa']] += ap['stake']
            if ap['cuentas'] in ('Ambas', 'Asenjo'):
                asenjo_apostado[ap['casa']] += ap['stake']

    yo_disp = {bm: round(max(0, yo.get(bm, 0) - yo_apostado.get(bm, 0)), 2) for bm in BOOKMAKERS}
    asenjo_disp = {bm: round(max(0, asenjo.get(bm, 0) - asenjo_apostado.get(bm, 0)), 2) for bm in BOOKMAKERS}

    yo_subtotal = round(sum(yo.get(k, 0) for k in SALDOS_COLS), 2)
    asenjo_subtotal = round(sum(asenjo.get(bm, 0) for bm in BOOKMAKERS), 2)

    return {
        'yo': yo,
        'asenjo': asenjo,
        'yo_apostado': dict(yo_apostado),
        'asenjo_apostado': dict(asenjo_apostado),
        'yo_disp': yo_disp,
        'asenjo_disp': asenjo_disp,
        'yo_subtotal': yo_subtotal,
        'asenjo_subtotal': asenjo_subtotal,
        'balance_total': round(yo_subtotal + asenjo_subtotal, 2),
        'yo_total_apostado': round(sum(yo_apostado.values()), 2),
        'asenjo_total_apostado': round(sum(asenjo_apostado.values()), 2),
    }


def _mes_sort_key(mes_str):
    try:
        abbr, yy = mes_str.split('-')
        return (2000 + int(yy), MONTH_NAMES.index(abbr) + 1)
    except (ValueError, AttributeError):
        return (0, 0)


def load_historial_legacy():
    """Meses anteriores a que existiera el registro mov-a-mov (hoja 'Historial',
    cargada a mano antes de este dashboard). No se recalculan, solo se exponen."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if 'Historial' not in wb.sheetnames:
        return []
    ws = wb['Historial']
    rows_iter = ws.iter_rows(min_row=2, values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    known_cols = set(BOOKMAKERS) | {'MES', 'BAL. APUESTAS', 'ENTRENADOR', 'CLASES', 'PAPIS', 'TRIBBU', 'BAL. TOTAL'}
    extra_bm_cols = [h for h in header if h and h not in known_cols]

    result = []
    for row in rows_iter:
        data = dict(zip(header, row))
        mes = data.get('MES')
        if not mes or not isinstance(mes, str):
            continue
        if not any(v not in (None, '') for k, v in data.items() if k != 'MES'):
            continue  # mes futuro sin datos aún
        row_out = {'mes': mes}
        bal_apuestas = 0.0
        for bm in BOOKMAKERS:
            v = _num(data.get(bm))
            row_out[bm] = round(v, 2)
            bal_apuestas += v
        for extra in extra_bm_cols:  # casas ya retiradas (p.ej. YAAS): se suman al total
            bal_apuestas += _num(data.get(extra))
        row_out['bal_apuestas'] = round(bal_apuestas, 2)
        row_out['clases'] = round(_num(data.get('CLASES')), 2)
        row_out['papis'] = round(_num(data.get('PAPIS')), 2)
        row_out['tribbu'] = round(_num(data.get('TRIBBU')), 2)
        row_out['entrenador'] = round(_num(data.get('ENTRENADOR')), 2)
        row_out['YAAS'] = round(_num(data.get('YAAS')), 2)
        row_out['bal_total'] = round(
            bal_apuestas + row_out['clases'] + row_out['papis'] + row_out['tribbu'] + row_out['entrenador'], 2)
        result.append(row_out)
    return result


def compute_historial(movimientos, apuestas):
    months = sorted({ap['mes'] for ap in apuestas if ap['mes']} |
                    {m['mes'] for m in movimientos if m['mes']})
    rows = []
    seen_meses = set()
    for mes_dt in months:
        mes_str = _fmt_mes(mes_dt)
        seen_meses.add(mes_str)
        row = {'mes': mes_str}
        bal_apuestas = 0.0
        for bm in BOOKMAKERS:
            gan = sum(ap['ganancia_yo'] + ap['ganancia_asenjo']
                      for ap in apuestas
                      if ap['mes'] == mes_dt and ap['casa'] == bm and not ap['pendiente'])
            row[bm] = round(gan, 2)
            bal_apuestas += gan
        yaas_gan = sum(ap['ganancia_yo'] + ap['ganancia_asenjo']
                       for ap in apuestas
                       if ap['mes'] == mes_dt and ap['casa'] == 'YAAS' and not ap['pendiente'])
        row['YAAS'] = round(yaas_gan, 2)
        bal_apuestas += yaas_gan
        row['bal_apuestas'] = round(bal_apuestas, 2)
        row['clases'] = round(sum(m['cantidad'] for m in movimientos
                                  if m['mes'] == mes_dt and m['tipo'] == 'Clases' and m['estado'] == 'Recibido'), 2)
        row['papis'] = round(sum(m['cantidad'] for m in movimientos
                                 if m['mes'] == mes_dt and m['tipo'] == 'Papis' and m['estado'] == 'Recibido'), 2)
        row['tribbu'] = round(sum(m['cantidad'] for m in movimientos
                                  if m['mes'] == mes_dt and m['tipo'] == 'Tribbu' and m['estado'] == 'Recibido'), 2)
        row['entrenador'] = round(sum(m['cantidad'] for m in movimientos
                                      if m['mes'] == mes_dt and m['tipo'] == 'Entrenador' and m['estado'] == 'Recibido'), 2)
        row['bal_total'] = round(bal_apuestas + row['clases'] + row['papis'] + row['tribbu'] + row['entrenador'], 2)
        rows.append(row)

    legacy_rows = [r for r in load_historial_legacy() if r['mes'] not in seen_meses]
    all_rows = legacy_rows + rows
    all_rows.sort(key=lambda r: _mes_sort_key(r['mes']))
    return all_rows


BOOKMAKERS_EXCLUSIVOS_CUENTA1 = {'BETFAIR', 'WILLIAMHILL', 'OLYBET', 'CASINOMAR BELLA'}


def _split_casa_mes(casa, mes_num, gan):
    """Reparto de la ganancia de una casa/mes entre Cuenta 1 y Cuenta 2. Marzo-26 va
    entero a Cuenta 1 (acordado a mano). El resto de meses: Betfair, Williamhill,
    Olybet y Casinomarbella van enteras a Cuenta 1; las demás casas se reparten al 50%."""
    if mes_num == 3:
        return gan, 0.0
    if casa in BOOKMAKERS_EXCLUSIVOS_CUENTA1:
        return gan, 0.0
    return round(gan / 2, 2), round(gan / 2, 2)


def compute_ganancias_por_cuenta(movimientos, apuestas, year=None):
    """Ganancias por cuenta en lo que va del año. Clases, Papis y Tribbu quedan fuera
    del cálculo; Yaas no tiene columna propia aquí (se muestra en el historial general)
    pero es 100% Cuenta 1, igual que Entrenador."""
    if year is None:
        year = datetime.now().year

    historial = compute_historial(movimientos, apuestas)
    rows = []
    for row in historial:
        anio, mes_num = _mes_sort_key(row['mes'])
        if anio != year:
            continue

        c1 = c2 = 0.0
        for bm in BOOKMAKERS:
            a, b = _split_casa_mes(bm, mes_num, row.get(bm, 0.0))
            c1 += a
            c2 += b
        c1 += row.get('YAAS', 0.0)
        entrenador = row.get('entrenador', 0.0)

        rows.append({
            'mes': row['mes'],
            'cuenta1_sin_entrenador': round(c1, 2),
            'cuenta1': round(c1 + entrenador, 2),
            'cuenta2': round(c2, 2),
            'total': round(c1 + entrenador + c2, 2),
        })
    return rows


def add_movimiento(fecha, mes_dt, tipo, cuenta, casa, cantidad, estado):
    _backup_before_write()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Movimientos']
    ws.append([fecha, mes_dt, tipo, cuenta, casa if casa else None, cantidad, estado])
    wb.save(EXCEL_PATH)


def _apuesta_ganancia_cells(casa, stake, cuota, resultado, cuentas):
    cuentas_v = _normalize_cuentas(casa, cuentas)
    gan = _calc_ganancia(resultado, stake, cuota)
    gan_yo, gan_asenjo = _split_ganancia(gan, cuentas_v)
    if gan_yo is None:
        return '', '', '', cuentas_v
    total = round(gan_yo + gan_asenjo, 2)
    return gan_yo, gan_asenjo, total, cuentas_v


def add_apuesta(mes_dt, fecha, partido, apuesta_desc, casa, stake, cuota, resultado, cuentas='Ambas'):
    _backup_before_write()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Registro_Apuestas']
    gan_yo, gan_asenjo, total, cuentas_v = _apuesta_ganancia_cells(casa, stake, cuota, resultado, cuentas)
    ws.append([
        mes_dt, fecha, partido, apuesta_desc, casa, stake, cuota, resultado,
        gan_yo, gan_asenjo, total, cuentas_v,
    ])
    wb.save(EXCEL_PATH)


def update_movimiento(row_num, fecha, mes_dt, tipo, cuenta, casa, cantidad, estado):
    _backup_before_write()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Movimientos']
    ws.cell(row=row_num, column=1).value = fecha
    ws.cell(row=row_num, column=2).value = mes_dt
    ws.cell(row=row_num, column=3).value = tipo
    ws.cell(row=row_num, column=4).value = cuenta
    ws.cell(row=row_num, column=5).value = casa if casa else None
    ws.cell(row=row_num, column=6).value = cantidad
    ws.cell(row=row_num, column=7).value = estado
    wb.save(EXCEL_PATH)


def update_movimiento_estado(row_num, estado):
    _backup_before_write()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Movimientos']
    ws.cell(row=row_num, column=7).value = estado
    wb.save(EXCEL_PATH)


def delete_movimiento(row_num):
    _backup_before_write()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Movimientos']
    ws.delete_rows(row_num, 1)
    wb.save(EXCEL_PATH)


def update_apuesta(row_num, mes_dt, fecha, partido, apuesta_desc, casa, stake, cuota, resultado, cuentas='Ambas'):
    _backup_before_write()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Registro_Apuestas']
    gan_yo, gan_asenjo, total, cuentas_v = _apuesta_ganancia_cells(casa, stake, cuota, resultado, cuentas)
    ws.cell(row=row_num, column=1).value = mes_dt
    ws.cell(row=row_num, column=2).value = fecha
    ws.cell(row=row_num, column=3).value = partido
    ws.cell(row=row_num, column=4).value = apuesta_desc
    ws.cell(row=row_num, column=5).value = casa
    ws.cell(row=row_num, column=6).value = stake
    ws.cell(row=row_num, column=7).value = cuota
    ws.cell(row=row_num, column=8).value = resultado
    ws.cell(row=row_num, column=9).value = gan_yo
    ws.cell(row=row_num, column=10).value = gan_asenjo
    ws.cell(row=row_num, column=11).value = total
    ws.cell(row=row_num, column=12).value = cuentas_v
    wb.save(EXCEL_PATH)


def update_apuesta_resultado(row_num, resultado):
    _backup_before_write()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Registro_Apuestas']
    casa = ws.cell(row=row_num, column=5).value
    stake = _num(ws.cell(row=row_num, column=6).value)
    cuota = _num(ws.cell(row=row_num, column=7).value)
    cuentas = ws.cell(row=row_num, column=12).value
    gan_yo, gan_asenjo, total, cuentas_v = _apuesta_ganancia_cells(casa, stake, cuota, resultado, cuentas)
    ws.cell(row=row_num, column=8).value = resultado
    ws.cell(row=row_num, column=9).value = gan_yo
    ws.cell(row=row_num, column=10).value = gan_asenjo
    ws.cell(row=row_num, column=11).value = total
    ws.cell(row=row_num, column=12).value = cuentas_v
    wb.save(EXCEL_PATH)


def delete_apuesta(row_num):
    _backup_before_write()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Registro_Apuestas']
    ws.delete_rows(row_num, 1)
    wb.save(EXCEL_PATH)


def get_available_months():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws_m = wb['Movimientos']
    ws_a = wb['Registro_Apuestas']
    months = set()
    for row in ws_m.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[1] and isinstance(row[1], datetime):
            months.add(_first_of_month(row[1]))
    for row in ws_a.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and isinstance(row[0], datetime):
            months.add(_first_of_month(row[0]))
    return sorted(months, reverse=True)
